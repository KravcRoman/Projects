import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

# Определение пути к файлу Excel
excel_file = 'Профиль_001.xlsx'
wb = load_workbook(excel_file, read_only=True)
ws = wb.active

# Поиск и извлечение данных расчётного периода
start_row = 10
end_row = 753
calculation_date = None
for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=2, max_col=2):
    for cell in row:
        try:
            calculation_date = datetime.strptime(cell.value, "%d.%m.%Y")
        except ValueError:
            pass
        if calculation_date:
            break
    if calculation_date:
        break

# Проверка наличия расчётного периода
if calculation_date is None:
    raise ValueError("Не удалось найти расчётный период в файле.")

# Извлечение месяца и года из даты
year = calculation_date.year
month = calculation_date.month

# Определение строки с номерами счетчиков и данных
meter_numbers_row = 5
for row in ws.iter_rows(min_row=meter_numbers_row, max_row=meter_numbers_row, min_col=4, max_col=ws.max_column):
    meter_numbers = [cell.value for cell in row if cell.value is not None]
    data_start_row = row[0].row + 1  # Показания начинаются со следующей строки

# Проверка, если номера счетчиков не были найдены
if not meter_numbers:
    raise ValueError("Номера счётчиков не найдены.")

# Проверка, если начальная строка данных не была определена
if not data_start_row:
    raise ValueError("Строка начала данных не определена.")


# Создание директорий для результатов
os.makedirs('wrk', exist_ok=True)
os.makedirs('errors', exist_ok=True)

# Функция для проверки данных
def check_data(meter_data):
    return not any([value is None for value in meter_data])

# Обработка данных для каждого счётчика
for meter_number in meter_numbers:
    meter_index = meter_numbers.index(meter_number) + 3  # Индекс счётчика в строке данных
    meter_data = [ws.cell(row=row_index, column=meter_index).value for row_index in range(data_start_row, end_row + 1) if ws.cell(row=row_index, column=meter_index).value is not None]
    if not check_data(meter_data):
        error_code = "Ошибка_данных"  # Пример кода ошибки
        filename = f'{meter_number}_{year}_{month:02d}_{error_code}.txt'
        filepath = os.path.join('errors', filename)
    else:
        filename = f'{meter_number}_{year}_{month:02d}_0.txt'
        filepath = os.path.join('wrk', filename)

    # Запись данных в текстовый файл
    with open(filepath, 'w') as f:
        for value in meter_data:
            f.write(f'{value},')
        f.write('\n')

print("Обработка завершена.")




